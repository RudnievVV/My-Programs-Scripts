#! python3
# zipcode_excel_match.py - opens Excel file with customers info and verifies customers's zipcode for existance and verifies itself with state.
# Works with predefined zipcodes in json file('USCities.json'), not with API requests in order to make it free and not depended to free 24-hours API requests limit.
# Works only with US zipcodes now, another zipcodes will be defined as INVALID ones (API enhancement will allow to work with worldwide zipcodes).
# IMPORTANT: it is needed to download 'USCities.json' file from the GitHub repository and put it inside working directory to let the program use it, link = https://github.com/RudnievVV/My-Programs-Scripts/blob/master/USCities.json
# Example of customers Excel file, link = https://github.com/RudnievVV/My-Programs-Scripts/blob/master/zipcode_excel_match_customers_example.xlsx


import openpyxl, json, re, os, sys, pymsgbox
from openpyxl.utils import get_column_letter, column_index_from_string


# Open Excel file and define cells for shipping and billing postal codes, state cells for shipping/billing
print('Excel file must contain headers for:\n1) shipping zipcode;\n2) billing zipcode;\n3) shipping state;\n4) billing state.\n\nOtherwise, program will not start.')
excel_file = input('Please enter the path to Excel file: ',)
while True:
        if os.path.isfile(excel_file) == True and \
           (
                   excel_file.endswith('.xlsx') or \
                   excel_file.endswith('.xlsm') or \
                   excel_file.endswith('.xltx') or \
                   excel_file.endswith('.xltm')
                   ):
                break
        else:
                excel_file = input('Please enter existing Excel file: ',)
excel_work_file = openpyxl.load_workbook(excel_file)
sheet = excel_work_file.active

bill_post_cell = None
ship_post_cell = None
bill_state_post_cell = None
ship_state_post_cell = None

for postal in range(1, sheet.max_column + 1):
        if 'postal' in sheet[get_column_letter(postal) + str('1')].value.lower() and 'bill' in sheet[get_column_letter(postal) + str('1')].value.lower():
                bill_post_cell = sheet[get_column_letter(postal) + str('1')]
        if 'postal' in sheet[get_column_letter(postal) + str('1')].value.lower() and 'ship' in sheet[get_column_letter(postal) + str('1')].value.lower():
                ship_post_cell = sheet[get_column_letter(postal) + str('1')]
        #if 'country' in sheet[get_column_letter(postal) + str('1')].value.lower() and 'bill' in sheet[get_column_letter(postal) + str('1')].value.lower():
                #bill_country_post_cell = sheet[get_column_letter(postal) + str('1')]
        #if 'country' in sheet[get_column_letter(postal) + str('1')].value.lower() and 'ship' in sheet[get_column_letter(postal) + str('1')].value.lower():
                #ship_country_post_cell = sheet[get_column_letter(postal) + str('1')]
        if 'state' in sheet[get_column_letter(postal) + str('1')].value.lower() and 'bill' in sheet[get_column_letter(postal) + str('1')].value.lower():
                bill_state_post_cell = sheet[get_column_letter(postal) + str('1')]
        if 'state' in sheet[get_column_letter(postal) + str('1')].value.lower() and 'ship' in sheet[get_column_letter(postal) + str('1')].value.lower():
                ship_state_post_cell = sheet[get_column_letter(postal) + str('1')]


# Check if all required headres are found, if not - quit the program
if bill_post_cell == None or \
   ship_post_cell == None or \
   bill_state_post_cell == None or \
   ship_state_post_cell == None:
        excel_work_file.close()
        print('One or more of the required headers are not found.\nNothing to check, program is closed.')
        sys.exit()


# Regex to define US/Canada/other zipcodes
regex = re.compile(r"(\w*)(\s|-)?(\w*)?")


# Open json file with zipcode data
with open('USCities.json') as us_zip_codes:
        json_us_zip_codes = json.load(us_zip_codes)

        # Check Excel file shipping and billing zipcodes for existance
        for cell in range(2, sheet.max_row + 1):
                succ_ship_state_count = None
                succ_bill_state_count = None
                correct_bill_zipcode_not_correct_bill_state = False
                correct_ship_zipcode_not_correct_ship_state = False
                print(f"Checking {cell} row...")
                
                if sheet[str(bill_post_cell.column_letter) + str(cell)].value == None or \
                   sheet[str(ship_post_cell.column_letter) + str(cell)].value == None:
                        print(f"Row {cell} is empty.")
                else:
                        for bill_code in range(len(json_us_zip_codes)):
                                if str(json_us_zip_codes[bill_code].get('zip_code')) == regex.search(str(sheet[str(bill_post_cell.column_letter) + str(cell)].value)).group(1) and \
                                   str(json_us_zip_codes[bill_code].get('state')) != str(sheet[str(bill_state_post_cell.column_letter) + str(cell)].value):
                                        correct_bill_zipcode_not_correct_bill_state = True
                                        print(f"Row {cell}, billing zipcode is OK, billing state is NOT OK!")
                                        
                                if str(json_us_zip_codes[bill_code].get('zip_code')) == regex.search(str(sheet[str(bill_post_cell.column_letter) + str(cell)].value)).group(1) and \
                                   str(json_us_zip_codes[bill_code].get('state')) == str(sheet[str(bill_state_post_cell.column_letter) + str(cell)].value):
                                        succ_bill_state_count = bill_code
                                        print(f"Row {cell}, billing zipcode is OK!")
                                        
                                if bill_code + 1 == len(json_us_zip_codes) and \
                                   succ_bill_state_count == None and \
                                   correct_bill_zipcode_not_correct_bill_state == False:
                                        print(f"!!!ROW {cell}, BILLING ZIPCODE IS INVALID!!!")
                                        
                        for ship_code in range(len(json_us_zip_codes)):
                                if str(json_us_zip_codes[ship_code].get('zip_code')) == regex.search(str(sheet[str(ship_post_cell.column_letter) + str(cell)].value)).group(1) and \
                                   str(json_us_zip_codes[ship_code].get('state')) != str(sheet[str(ship_state_post_cell.column_letter) + str(cell)].value):
                                        correct_ship_zipcode_not_correct_ship_state = True
                                        print(f"Row {cell}, shipping zipcode is OK, shipping state is NOT OK!")
                                        
                                if str(json_us_zip_codes[ship_code].get('zip_code')) == regex.search(str(sheet[str(ship_post_cell.column_letter) + str(cell)].value)).group(1) and \
                                   str(json_us_zip_codes[ship_code].get('state')) == str(sheet[str(ship_state_post_cell.column_letter) + str(cell)].value):
                                        succ_ship_state_count = ship_code
                                        print(f"Row {cell}, shipping zipcode is OK!")
                                        
                                if ship_code + 1 == len(json_us_zip_codes) and \
                                   succ_ship_state_count == None and \
                                   correct_ship_zipcode_not_correct_ship_state == False:
                                        print(f"!!!ROW {cell}, SHIPPING ZIPCODE IS INVALID!!!")


print('---------\nDone.')
excel_work_file.close()
pymsgbox.alert(text='Excel zipcode file checking was finished successfully.', \
               title='Excel zipcode checking')
sys.exit()
